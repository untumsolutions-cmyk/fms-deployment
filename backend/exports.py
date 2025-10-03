from fastapi import APIRouter, HTTPException, Response, Query
from pathlib import Path
import sqlite3, os, io
try:
    import openpyxl
except Exception:
    openpyxl = None
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet
except Exception:
    raise

router = APIRouter()
BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = BASE_DIR / 'templates'
UPLOADS = BASE_DIR / 'uploads' / 'logos'

def get_db():
    conn = sqlite3.connect(BASE_DIR / 'fms.db')
    conn.row_factory = sqlite3.Row
    return conn

def logo_path_for(filename):
    p = UPLOADS / filename
    return p if p.exists() else None

def make_pdf_bytes(title, header_lines, table_rows, logo_file=None):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30,leftMargin=30, topMargin=30,bottomMargin=18)
    elements = []
    styles = getSampleStyleSheet()
    if logo_file:
        try:
            img = Image(str(logo_file))
            img.drawHeight = 50
            img.drawWidth = 100
            elements.append(img)
        except Exception:
            pass
    elements.append(Paragraph(f'<b>{title}</b>', styles['Title']))
    elements.append(Spacer(1,12))
    for line in header_lines:
        elements.append(Paragraph(line, styles['Normal']))
    elements.append(Spacer(1,12))
    if table_rows:
        # header
        header = list(table_rows[0].keys())
        data = [header]
        for r in table_rows:
            data.append([str(r.get(h,'')) for h in header])
    else:
        data = [['No data']]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0.5,colors.black),('BACKGROUND',(0,0),(-1,0),colors.grey),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke)]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer.read()

@router.get('/export/invoice/{invoice_id}')
def export_invoice(invoice_id: int, format: str = Query('pdf', enum=['pdf','xlsx']), logo: str = None):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM invoices WHERE invoice_id = ?', (invoice_id,))
    inv = cur.fetchone()
    if not inv:
        raise HTTPException(status_code=404, detail='Invoice not found')
    cur.execute('SELECT * FROM invoice_items WHERE invoice_id = ?', (invoice_id,))
    items = [dict(r) for r in cur.fetchall()]
    # build xlsx if template exists
    xlsx_bytes = None
    if openpyxl is not None:
        candidates = list(TEMPLATE_DIR.glob('*invoice*.xlsx')) + list(TEMPLATE_DIR.glob('*Invoice*.xlsx')) + list(TEMPLATE_DIR.glob('*service*.xlsx')) + list(TEMPLATE_DIR.glob('*trading*.xlsx'))
        if candidates:
            tpl = candidates[0]
            wb = openpyxl.load_workbook(tpl)
            ws = wb.active
            try:
                ws['B1'] = inv['invoice_id']
                ws['B2'] = inv['customer_id']
                ws['B3'] = inv['date']
            except Exception:
                pass
            start_row = 10
            for i, it in enumerate(items):
                r = start_row + i
                ws.cell(row=r, column=1, value=it.get('description'))
                ws.cell(row=r, column=2, value=it.get('quantity'))
                ws.cell(row=r, column=3, value=it.get('unit_price'))
                ws.cell(row=r, column=4, value=it.get('total'))
            out = io.BytesIO()
            wb.save(out)
            xlsx_bytes = out.getvalue()
    # build PDF via reportlab (universal)
    logo_file = logo_path_for(logo) if logo else None
    header_lines = [f'Date: {inv["date"]}', f'Customer ID: {inv["customer_id"]}']
    pdf = make_pdf_bytes(f'Invoice #{inv["invoice_id"]}', header_lines, items, logo_file=logo_file)
    if format == 'xlsx' and xlsx_bytes is not None:
        return Response(content=xlsx_bytes, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition':f'attachment; filename=invoice-{invoice_id}.xlsx'})
    return Response(content=pdf, media_type='application/pdf', headers={'Content-Disposition':f'attachment; filename=invoice-{invoice_id}.pdf'})

# other exports: quote, statement, monthly, ageing, balances, payslip
@router.get('/export/quote/{quote_id}')
def export_quote(quote_id:int, format: str = Query('pdf', enum=['pdf','xlsx']), logo: str = None):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM quotes WHERE quote_id = ?', (quote_id,))
    q = cur.fetchone()
    if not q:
        raise HTTPException(status_code=404, detail='Quote not found')
    rows = [dict(q)]
    pdf = make_pdf_bytes(f'Quote #{quote_id}', [], rows, logo_file=(logo_path_for(logo) if logo else None))
    return Response(content=pdf, media_type='application/pdf', headers={'Content-Disposition':f'attachment; filename=quote-{quote_id}.pdf'})

@router.get('/export/statement/{customer_id}')
def export_statement(customer_id:int, format: str = Query('pdf', enum=['pdf','xlsx']), logo: str = None):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM transactions WHERE customer_id = ?', (customer_id,))
    rows = [dict(r) for r in cur.fetchall()]
    pdf = make_pdf_bytes(f'Statement for Customer {customer_id}', [], rows, logo_file=(logo_path_for(logo) if logo else None))
    return Response(content=pdf, media_type='application/pdf', headers={'Content-Disposition':f'attachment; filename=statement-{customer_id}.pdf'})

@router.get('/export/monthly/{customer_id}')
def export_monthly(customer_id:int, logo: str = None):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT date, SUM(amount) as total FROM transactions WHERE customer_id = ? GROUP BY date', (customer_id,))
    rows = [dict(r) for r in cur.fetchall()]
    pdf = make_pdf_bytes(f'Monthly summary for {customer_id}', [], rows, logo_file=(logo_path_for(logo) if logo else None))
    return Response(content=pdf, media_type='application/pdf', headers={'Content-Disposition':f'attachment; filename=monthly-{customer_id}.pdf'})

@router.get('/export/ageing/{customer_id}')
def export_ageing(customer_id:int, logo: str = None):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT invoice_id, date, balance_due FROM invoices WHERE customer_id = ? AND balance_due > 0', (customer_id,))
    rows = [dict(r) for r in cur.fetchall()]
    pdf = make_pdf_bytes(f'Ageing for {customer_id}', [], rows, logo_file=(logo_path_for(logo) if logo else None))
    return Response(content=pdf, media_type='application/pdf', headers={'Content-Disposition':f'attachment; filename=ageing-{customer_id}.pdf'})

@router.get('/export/balances/{customer_id}')
def export_balances(customer_id:int, logo: str = None):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT account_name, account_type FROM accounts')
    rows = [dict(r) for r in cur.fetchall()]
    pdf = make_pdf_bytes(f'Balances', [], rows, logo_file=(logo_path_for(logo) if logo else None))
    return Response(content=pdf, media_type='application/pdf', headers={'Content-Disposition':f'attachment; filename=balances-{customer_id}.pdf'})

@router.get('/export/payslip/{payslip_id}')
def export_payslip(payslip_id:int, format: str = Query('pdf', enum=['pdf','xlsx']), logo: str = None):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM payslips WHERE payslip_id = ?', (payslip_id,))
    p = cur.fetchone()
    if not p:
        raise HTTPException(status_code=404, detail='Payslip not found')
    rows = [dict(p)]
    # xlsx generation if template exists
    xlsx_bytes = None
    if openpyxl is not None:
        tpl = TEMPLATE_DIR / 'payslip_template.xlsx'
        if tpl.exists():
            wb = openpyxl.load_workbook(tpl)
            ws = wb.active
            try:
                ws['B2'] = p['payslip_id']
                ws['B3'] = p['employee_id']
                ws['B4'] = p['period_start']
                ws['B5'] = p['period_end']
                ws['B6'] = p['gross_salary']
                ws['B7'] = p['total_deductions']
                ws['B8'] = p['net_salary']
            except Exception:
                pass
            out = io.BytesIO()
            wb.save(out)
            xlsx_bytes = out.getvalue()
    pdf = make_pdf_bytes(f'Payslip #{p["payslip_id"]}', [], rows, logo_file=(logo_path_for(logo) if logo else None))
    if format == 'xlsx' and xlsx_bytes is not None:
        return Response(content=xlsx_bytes, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition':f'attachment; filename=payslip-{payslip_id}.xlsx'})
    return Response(content=pdf, media_type='application/pdf', headers={'Content-Disposition':f'attachment; filename=payslip-{payslip_id}.pdf'})
